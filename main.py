import tkinter as tk
from tkinter import ttk
import openpyxl,xlwings as xw

# edit button updates treeview/sheet
def edit_row():
    zeta=treeview.selection()
    temp=treeview.item(zeta,'values')
    if zeta:
        new_paid=amt_entry.get()
        new_status=status_combobox.get()
        # print(new_paid,new_status)
        if new_status and new_paid:
            # treeview.item(zeta,values=(temp[0],temp[1],temp[2],temp[3],new_status,temp[5],temp[6],new_paid))
            find_row(temp[1],new_paid,new_status)

            data = [(treeview.set(item,"Amount Due"), item) for item in treeview.get_children()]
            data.sort(reverse=False)
            for index, (val, item) in enumerate(data):
                treeview.move(item, '', index)

 

    # name=name_entry.get()
    # acc=int(acc_entry.get())
    # amt=int(amt_entry.get())
    # stat=status_combobox.get
    # zeta=treeview.selection()
    # if zeta:
    #     values=treeview.item(zeta)['values']
    #     if values:
    #         if not (status_combobox.state(["Not Paid"])):
    #             amtfinal=int(values[2])-amt
    #             print(amtfinal)

def find_row(value,new_paid,new_status):
    workbook = openpyxl.load_workbook("t1.xlsx")
    sheet = workbook.active
    target_value = int(value)
    # Initialize a variable to store the row number if the value is found
    row_number = None
    for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=0):
        cell = row[0]  # Access the cell in the target column
        cell_value = cell.value
        if cell_value == target_value:
            row_number = cell.row
            break # Stop the loop once a match is found
    workbook.close()

    # using xlwings
    app = xw.App(visible=False)
    workbook1=app.books.open('t1.xlsx')
    sheet1=workbook1.sheets['Sheet1']
    #modifying values into workbook
    sheet1[f'F{row_number}'].value=new_status
    if sheet1[f'H{row_number}'].value==None:
        sheet1[f'H{row_number}'].value=int(new_paid)
    else:
        sheet1[f'H{row_number}'].value+=int(new_paid)

    workbook1.save('t1.xlsx')
    workbook1.save('t2.xlsx')   #optional

    # for row in sheet.used_range.rows[2:]:
    #       # Skip the header row
    #     values = [cell.value for cell in row]
    #     treeview.insert('', tk.END, values=(values[2], values[1], values[7], values[5], values[6], values[4], values[3]))
    workbook1.close()
    app.quit()
    
    wb=openpyxl.load_workbook("t1.xlsx",data_only=True)
    sheet=wb.active
    for item in treeview.get_children():
        treeview.delete(item)

    for row in sheet.iter_rows(min_row=2, values_only=True):#decremented row by 1,to revert simply add 1
        if row[0]==None:
            break
        treeview.insert('', tk.END, values=(row[1], row[0], row[6], row[4], row[5], row[3], row[2]))
    wb.close()


# toggling themes
def toggle_mode():
    if mode_switch.instate(["selected"]):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")


# initial loading data into treeview 

# def load_data(): this is for a.xlsx table,testing small sample file
#     specific_indexes=[0,1,6,2,3,4]
#     wb=openpyxl.load_workbook("a.xlsx",data_only=True)
#     sheet=wb.active
#     # obtain all data from xl as a list
#     list_val=list(sheet.values)
#     #/ print(list_val)
#     heading=["Name","Account No","Amount Due","Due by","Status"]
#     for col in heading:
#         treeview.column(column=col,anchor="center")
#         treeview.heading(col,text=col)
#     for val in list_val[1:]:                 
#         treeview.insert('', tk.END, values=(val[0],val[1],val[6],val[3],val[4],val[5]))
#         # treeview.insert('', tk.END, values=val,)

# def load_data(): #this is for t1.xlsx pretty much close to similar data file to operate on
#     specific_indexes=[0,1,6,2,3,4]
#     wb=openpyxl.load_workbook("t1.xlsx",data_only=True)
#     sheet=wb.active
#     # obtain all data from xl as a list
#     list_val=list(sheet.values)
#     #/ print(list_val)
#     heading=["Name","Account No","Amount Due","Due by","Status"]#val[4] is months_paid,val[3] is denomination
#     for col in heading:
#         treeview.column(column=col)
#         treeview.heading(col,text=col)
#     for val in list_val[1:]:           #values[0]     [1]    [2]    [3]    [4]    [5]    [6]            
#         treeview.insert('', tk.END, values=(val[2],val[1],val[7],val[5],val[6],val[4],val[3]))
#         # treeview.insert('', tk.END, values=val,)
#     wb.close()

def load_data(): #this is for t1.xlsx pretty much close to similar data file to operate on
    specific_indexes=[0,1,6,2,3,4]
    wb=openpyxl.load_workbook("t1.xlsx",data_only=True)
    sheet=wb.active
    # obtain all data from xl as a list
    list_val=list(sheet.values)
    #/ print(list_val)
    heading=["Name","Account No","Amount Due","Due by","Status"]#val[4] is months_paid,val[3] is denomination
    for col in heading:
        treeview.column(column=col)
        treeview.heading(col,text=col) #decremented val by 1,to revert simply add 1
    for val in list_val[1:]:           #values[0]     [1]    [2]    [3]    [4]    [5]    [6]
        if val[1]==None:
            break            
        treeview.insert('', tk.END, values=(val[1],val[0],val[6],val[4],val[5],val[3],val[2]))
        # treeview.insert('', tk.END, values=val,)
    wb.close()

# loading data to entry fields upon selection of treeview
def extract_data(a):
    # zeta has selected treeview information such as {target,['values'] etc} and values contain ['values'] data
    zeta=treeview.selection()
    if zeta:
        values=treeview.item(zeta)['values']
        if values:
            name_entry.config(state="active")
            name_entry.delete(0,"end")
            name_entry.insert(0,values[0])
            name_entry.config(state="readonly")

            acc_entry.config(state="active")
            acc_entry.delete(0,"end")
            acc_entry.insert(0,values[1])
            acc_entry.config(state="readonly")
            
            amtthere_entry.config(state="active")
            amtthere_entry.delete(13,"end")
            amtthere_entry.insert("end",values[6])
            amtthere_entry.config(state="readonly")

            amt_entry.delete(0,"end")
            amt_entry.insert(0,values[2])

            status_combobox.delete(0,"end")
            status_combobox.insert(0,values[4])

            month_entry.config(state='active')
            month_entry.delete(0,"end")
            month_entry.insert(0,values[5])
            month_entry.config(state='readonly')

    
    #    selectedItem = trv.selection()[0]
    # entry_id.insert(0, trv.item(selectedItem)['values'][0])
    # entry_fname.insert(0, trv.item(selectedItem)['values'][1])
    # entry_lname.insert(0, trv.item(selectedItem)['values'][2])
    # entry_email.insert(0, trv.item(selectedItem)['values'][3])
    # entry_age.insert(0, trv.item(selectedItem)['values'][4])

root=tk.Tk()                                                                                             # root frame as base for all frame
style=ttk.Style(root)                                                                                    #theming purpose                    
root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")
frame=ttk.Frame(root)                                                                                    #frame on root,pack() to arrange more widgets efficiently(similar like grid,place)
frame.pack(expand=True)

widgets_frame = ttk.LabelFrame(frame, text="Account Details")                                            #widget_frame on frame for various entry fields like name,acc,amt
widgets_frame.grid(row=0,column=0)

# a=tk.BooleanVar()                                                                                        #check to unlock editing the fields
# checkbutton = ttk.Checkbutton(widgets_frame, text="Edit", variable=a)
# checkbutton.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

name_entry=ttk.Entry(widgets_frame)
name_entry.grid(row=0,column=0,padx=5,pady=4, sticky="ew")
name_entry.insert(0,"Name")
name_entry.bind("<FocusIn>",lambda e:name_entry.delete(0,"end"))

acc_entry=ttk.Entry(widgets_frame)
acc_entry.grid(row=1,column=0,padx=5,pady=4, sticky="ew")
acc_entry.insert(0,"Account number")
acc_entry.bind("<FocusIn>",lambda e:acc_entry.delete(0,"end"))

amtthere_entry=ttk.Entry(widgets_frame)
amtthere_entry.grid(row=2,column=0,padx=5,pady=4, sticky="ew")
amtthere_entry.insert(0,"Denomination:")
amtthere_entry.config(state="readonly")

amt_entry=ttk.Entry(widgets_frame)
amt_entry.grid(row=3,column=0,padx=5,pady=4, sticky="ew")
amt_entry.insert(0,"Amount Paid")
amt_entry.bind("<FocusIn>",lambda e:amt_entry.delete(0,"end"))

combo_list=["Not Paid","Paid-Cash","Paid-GPay","Paid-PhonePe","Paid-Net Banking"]
status_combobox=ttk.Combobox(widgets_frame,values=combo_list)
status_combobox.grid(row=4,column=0,padx=5,pady=4, sticky="ew")
status_combobox.current(0)

month_entry=ttk.Entry(widgets_frame,)
month_entry.grid(row=5,column=0,padx=5,pady=4, sticky="ew")
month_entry.insert(0,"Installment Remaining")
month_entry.bind("<FocusIn>",lambda e:month_entry.delete(0,"end"))

button=ttk.Button(widgets_frame,text="Submit",command=edit_row)               
button.grid(row=6,column=0,pady=(15,10))

separator = ttk.Separator(widgets_frame)
separator.grid(row=7, column=0, padx=(10, 10), pady=7, sticky="ew")

mode_switch = ttk.Checkbutton(widgets_frame, text="Mode", style="Switch", command=toggle_mode)
mode_switch.grid(row=8, column=0, padx=10, pady=10, sticky="nsew")

# separator2=ttk.Separator(widgets_frame)
# separator2.grid(row=9,column=0.pad)
separator2 = ttk.Separator(frame)
separator2.grid(row=0, column=1, padx=(10, 10), pady=(15,8), sticky="ns")


treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=2, pady=10,)

cols=("Name","Account No","Amount Due","Due by","Status")

treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="both")
treeview = ttk.Treeview(treeFrame, show="headings",
                        yscrollcommand=treeScroll.set, columns=cols, height=15,)

treeview.column("Name", width=275)
treeview.column("Account No", width=110,anchor='center')
treeview.column("Amount Due", width=100)
treeview.column("Due by", width=70)
treeview.column("Status", width=120,anchor='center')
treeview.bind("<<TreeviewSelect>>",extract_data)
# treeview.bind("<Button-2>",lambda e:treeview.config(height=+30,))
treeview.pack(pady=(10,5),fill='x',expand=True)
treeScroll.config(command=treeview.yview)
load_data()

def search(a):
    query = search_entry.get().lower()  # Get the search query and convert to lowercase for case-insensitive search
    for item in treeview.get_children():
        values = treeview.item(item, 'values')
        if query in [str(value).lower() for value in values]:
            for item in treeview.get_children():
                treeview.delete(item)
            treeview.insert("", "end",values=values)

def clear_search(a):
    search_entry.delete(0, tk.END)  # Clear the search entry
    
    # Remove all items from the Treeview and insert all items
    for item in treeview.get_children():
        treeview.delete(item)
    load_data()

search_entry = ttk.Entry(frame,width=30)
search_entry.grid(row=2,column=2,pady=(0,10),padx=15,sticky='e')
search_entry.bind("<Return>",search)
search_entry.bind("<MouseWheel>",clear_search)


root.mainloop()                                                                                         #https://www.youtube.com/watch?v=8m4uDS_nyCk source